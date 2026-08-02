"""
Chemical import-substitution HSN-8 signal analysis.

Inputs: DGCI&S TradeStat commodity-wise Import/Export Value files (US$M) and
the companion Import Quantity file, all at the 8-digit HS code level.

Produces, per qualifying HSN-8 code (FY26 import - FY26 export >= $0.5M):
  - value/volume/price growth FY24-25 -> FY25-26
  - a signal classification against the import-substitution framework:
      A_SCISSORS          value & volume both falling, price rising      -> active substitution
      B_INELASTIC_TARGET  top-quartile value, price flat, value steady   -> PLI/capex candidate
      D_DEEPENING         value & volume both rising fast                -> dependency growing
      E_NO_CLEAR_SIGNAL   no clear 1-FY directional pattern
      C_INSUFFICIENT_DATA no matched quantity row
  - a rule-based qualitative rationale tag (fertiliser/petrochem/agrochem/etc.)

Known limits (see the published article's methodology blockquote for the
user-facing version of this list):
  - only a 1-FY window, not the 3-5yr trend the framework calls for
  - no Import Penetration Ratio (needs domestic production data, not in TradeStat)
  - no Component Shift (Signal C) detection (needs finished-vs-component HSN pairing)
  - no country-of-origin / FTA-rerouting check (TradeStat commodity-wise has no partner country)
  - no systematic tariff/PLI cross-reference per HSN line (only qualitative keyword tags)
"""
import openpyxl
import json
import re

CHEM_CHAPTERS = ['28', '29', '31', '32', '33', '34', '38', '39']

CH_NAMES = {
    '28': 'Inorganic Chemicals & Fertiliser Intermediates',
    '29': 'Organic Chemicals',
    '31': 'Fertilisers',
    '32': 'Dyes, Pigments, Tanning & Colouring Extracts',
    '33': 'Essential Oils, Cosmetics & Perfumery',
    '34': 'Soaps, Waxes, Polishes & Surface-Active Agents',
    '38': 'Miscellaneous Chemical Products',
    '39': 'Plastics & Articles Thereof',
}

RATIONALE_RULES = [
    (r'\b(UREA|DAP|DIAMONM|MURIATE|POTASH|POTASSIUM CHLORIDE|SUPERPHOSPHAT|ROCK PHOSPHATE|SULPHATE OF POTASH|NPK|FERTLSR|FERTILIS)\b',
     'Fertiliser feedstock/finished — India lacks phosphate rock & potash reserves; imports are structural, not import-substitutable without mining tie-ups (Morocco/Jordan/Canada/Belarus).'),
    (r'\b(GOLD|PLATINUM|RHODIUM|PALLADIUM|SILVER|NOBLE METAL)\b',
     'Precious/noble metal — natural-resource constrained; India has negligible reserves, substitution means recycling/urban-mining not manufacturing.'),
    (r'\b(POLYPROPYLENE|POLYETHYLENE|PVC|POLYCARBONAT|POLYMER|COPOLYMER|RESIN|ETHYLENE|PROPYLENE)\b',
     'Base petrochemical/polymer — cracker capacity gap; substitutable via new naphtha/gas cracker + downstream polymer trains (PLI/PCPIR candidate).'),
    (r'\b(BENZENE|STYRENE|TOLUENE|XYLENE|METHANOL|ALCOHOL|ACETIC|FORMALDEHYDE|ACRYLIC|PHENOL)\b',
     'Basic organic intermediate — aromatics/methanol value chain; substitutable with refinery-integrated petrochemical capacity.'),
    (r'\b(INSECTICID|FUNGICID|HERBICID|PESTICID|AGROCHEM|CARTAP|MANCOZEB)\b',
     'Agrochemical technical/formulation — patent-expired generics; substitutable via technical-grade manufacturing (China+1 opportunity).'),
    (r'\b(DYE|PIGMENT|COLOUR|COLORANT|PHATHALOCYANINE|PATHALOCYANINE)\b',
     'Specialty dye/pigment — process-technology & environmental-compliance gap (dye intermediates are pollution-intensive); substitutable with clean-tech investment.'),
    (r'\b(VITAMIN|ANTIBIOTIC|BULK DRUG|API|PHARMA|STEROID|HORMONE)\b',
     'Pharma API/intermediate — precursor-chemistry dependency on China; strategic substitution priority (PLI-for-APIs scheme already targets this).'),
    (r'\b(ACID)\b',
     'Industrial acid — commodity intermediate; substitutable via capacity addition, often import-parity priced so investment case is thin unless anti-dumping applies.'),
    (r'\b(OIL|ESSENTIAL OIL|PERFUM|FRAGRANC|AROMATIC)\b',
     'Essential oil/fragrance compound — agro-botanical or specialty-synthesis input; partly substitutable via natural-farming linkages (lavender/lemongrass missions).'),
    (r'\b(ENZYME|CATALYST|REAGENT|DIAGNOSTIC)\b',
     'Catalyst/enzyme/diagnostic reagent — high-IP specialty chemistry; substitution needs licensed tech transfer, not commodity capacity.'),
    (r'\b(WAX|SOAP|SURFACE|SURFACTANT|POLISH)\b',
     'Surfactant/specialty formulation — oleochemical feedstock or IP-licensed formulation; moderate substitution potential via oleochemical integration.'),
    (r'\b(RUBBER|LATEX)\b',
     'Rubber chemical/latex-linked input — natural rubber deficit compounds the synthetic-chemical gap.'),
    (r'\b(CARBON BLACK|ACTIVATED CARBON)\b',
     'Carbon-based industrial input — feedstock (CBFS/coal) linked; substitutable with refinery-integrated carbon black capacity.'),
]
RATIONALE_DEFAULT = 'Chemical intermediate/speciality — deficit driven by feedstock, scale-economics or technology gap; needs case-by-case investment screening.'


def classify_rationale(name):
    for pat, why in RATIONALE_RULES:
        if re.search(pat, name, re.I):
            return why
    return RATIONALE_DEFAULT


def load_value_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[1]:
            continue
        hs = str(r[1]).strip()
        if len(hs) != 8:
            continue
        rows.append({'hs': hs, 'name': (r[2] or '').strip(), 'fy25': r[3] or 0, 'fy26': r[5] or 0, 'growth': r[7]})
    return rows


def load_qty_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[1]:
            continue
        hs = str(r[1]).strip()
        if len(hs) != 8:
            continue
        rows[hs] = {'unit': r[3], 'qty25': r[4] or 0, 'qty26': r[5] or 0, 'qty_growth': r[6]}
    return rows


def classify_signal(o, p75_value):
    if o['imp_fy25']:
        vg = (o['imp_fy26'] - o['imp_fy25']) / o['imp_fy25'] * 100
    else:
        vg = None
    qg = o.get('qty_growth_pct')
    pg = o.get('price_growth_pct')

    if vg is None or qg is None or pg is None:
        return 'C_INSUFFICIENT_DATA'
    if vg <= -10 and qg <= -10 and pg >= 5:
        return 'A_SCISSORS'
    if o['imp_fy26'] >= p75_value and abs(pg) <= 8 and vg > -5:
        return 'B_INELASTIC_TARGET'
    if vg >= 15 and qg >= 15:
        return 'D_DEEPENING'
    return 'E_NO_CLEAR_SIGNAL'


def build(import_value_xlsx, export_value_xlsx, import_qty_xlsx, deficit_floor_musd=0.5):
    imp = load_value_file(import_value_xlsx)
    exp = load_value_file(export_value_xlsx)
    imp_qty = load_qty_file(import_qty_xlsx)
    exp_map = {r['hs']: r for r in exp}

    qualifying = []
    for r in imp:
        ch = r['hs'][:2]
        if ch not in CHEM_CHAPTERS:
            continue
        e = exp_map.get(r['hs'], {'fy26': 0})
        deficit = r['fy26'] - e.get('fy26', 0)
        if r['fy26'] < 1 or deficit < deficit_floor_musd:
            continue
        o = {
            'hs': r['hs'], 'chapter': ch, 'name': r['name'],
            'imp_fy25': r['fy25'], 'imp_fy26': r['fy26'], 'imp_growth': r['growth'],
            'exp_fy26': e.get('fy26', 0), 'deficit_fy26': round(deficit, 1),
            'rationale': classify_rationale(r['name']),
        }
        q = imp_qty.get(r['hs'])
        if q and q['qty25'] and q['qty26']:
            up25 = r['fy25'] / q['qty25']
            up26 = r['fy26'] / q['qty26']
            o.update({
                'qty25': q['qty25'], 'qty26': q['qty26'], 'unit': q['unit'],
                'qty_growth_pct': q['qty_growth'],
                'unit_price_fy25': round(up25, 4), 'unit_price_fy26': round(up26, 4),
                'price_growth_pct': round((up26 - up25) / up25 * 100, 1) if up25 else None,
            })
        else:
            o.update({'qty25': None, 'qty26': None, 'unit': None, 'qty_growth_pct': None,
                      'unit_price_fy25': None, 'unit_price_fy26': None, 'price_growth_pct': None})
        qualifying.append(o)

    vals = sorted(o['imp_fy26'] for o in qualifying)
    p75 = vals[int(len(vals) * 0.75)] if vals else 0
    for o in qualifying:
        o['is_signal'] = classify_signal(o, p75)

    by_chapter = {}
    for o in qualifying:
        by_chapter.setdefault(o['chapter'], []).append(o)

    return {'by_chapter': by_chapter, 'ch_names': CH_NAMES, 'p75_fy26': p75}


if __name__ == '__main__':
    import sys
    downloads = '/Users/umashankar/Downloads/'
    result = build(
        downloads + 'TradeStat-Eidb-Import-Commodity-wise.xlsx',
        downloads + 'TradeStat-Eidb-Export-Commodity-wise.xlsx',
        downloads + 'TradeStat-Eidb-Import-Commodity-wise (1).xlsx',
    )
    out_path = sys.argv[1] if len(sys.argv) > 1 else 'chem_hsn8_signals.json'
    json.dump(result, open(out_path, 'w'))
    total = sum(len(v) for v in result['by_chapter'].values())
    print(f'{total} qualifying codes written to {out_path}')
