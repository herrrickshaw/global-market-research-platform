#!/usr/bin/env python3
"""
scan_core.py — the ONE implementation of scan logic shared by all five market scans.

WHY (2026-07-27, SCREENER_SMOOTHNESS_STRATEGIES #1): the knowledge graph found
compute_darvas_box() in 13 copies; a source diff then showed they had ALREADY
drifted the dangerous way — JP/KR/EU carried the trailing-NaN fix (Korea
2026-07-21: .fillna(0) marked 2,472/2,480 stocks BREAKDOWN_SELL) while IN/US
never received it. This module unifies on the FIXED semantics and parameterizes
the real per-market differences (column names, price rounding).

The five full_*_market_scan.py keep their public function names as thin
wrappers around this module, so no call site changes.

Design rules preserved (see feedback memory + inline history):
  • current bar is EXCLUDED from box formation (breakdown detection dies otherwise)
  • trailing NaN bars are DROPPED, never zero-filled (zero is a price, not a null)
  • a missing final close is NO_DATA, never a silent IN_BOX
"""
from __future__ import annotations

import pandas as pd

DARVAS_CONFIRM = 3   # consecutive bars a high/low must hold to confirm a box edge


# ── Darvas box ────────────────────────────────────────────────────────────────
def compute_darvas_box(df: pd.DataFrame, confirm: int = DARVAS_CONFIRM,
                       decimals: int = 2, col_candidates: dict | None = None) -> dict:
    """Detect Darvas Box and classify current price.

    decimals: price rounding (2 for ₹/$/€, 0 for ¥/₩ whole-unit quotes).
    col_candidates: optional {"high": [...], "low": [...], "close": [...]}
    name-candidates for exotic schemas (India bhavcopy CH_* columns).

    Returns a SUPERSET of the historical per-market schemas so every existing
    caller keeps working: both `pos_in_box`/`position_in_box_pct` and
    `upside_pct`/`upside_to_top_pct` spellings, plus `box_range`.
    """
    cand = col_candidates or {}

    def find_col(frame, names):
        for c in names:
            m = next((col for col in frame.columns if c.upper() in col.upper()), None)
            if m:
                return m
        return None

    if df is None or df.empty:
        return {"signal": "INSUFFICIENT_DATA", "box_top": None, "box_bottom": None}

    h_col = find_col(df, cand.get("high", ["High"]))
    l_col = find_col(df, cand.get("low", ["Low"]))
    c_col = find_col(df, cand.get("close", ["Close"]))
    if not all([h_col, l_col, c_col]) or len(df) < confirm + 5:
        return {"signal": "INSUFFICIENT_DATA", "box_top": None, "box_bottom": None}

    # Drop trailing bars with no close BEFORE anything else. pykrx/yfinance
    # append a row for the current session as soon as the date exists, so a scan
    # run before the market prints sees a NaN final bar. The correct "current
    # price" is the last SETTLED close — never a null and certainly not zero.
    # 🔴 NaN bars must be DROPPED, not zero-filled: .fillna(0) made current=0,
    # below every box bottom → Korea 2026-07-21 reported BREAKDOWN_SELL on
    # 2,472 of 2,480 rows with Position_in_Box% at -1990.9. Zero is a PRICE.
    try:
        _c = pd.to_numeric(df[c_col], errors="coerce")
        _last = _c.last_valid_index()
        if _last is None:
            return {"signal": "NO_DATA", "box_top": None, "box_bottom": None,
                    "current_price": None}
        df = df.loc[:_last]
    except Exception:
        pass

    highs = pd.to_numeric(df[h_col], errors="coerce").tolist()
    lows = pd.to_numeric(df[l_col], errors="coerce").tolist()
    closes = pd.to_numeric(df[c_col], errors="coerce").tolist()

    current = closes[-1]
    if current is None or current != current or current <= 0:
        # Explicit: a missing final bar is NO_DATA. Falling through would make
        # every comparison False and silently report IN_BOX — a different
        # wrong answer, not a right one.
        return {"signal": "NO_DATA", "box_top": None, "box_bottom": None,
                "current_price": None}

    h = highs[:-1]   # exclude current bar from box formation (design rule)
    l = lows[:-1]
    n = len(h)

    box_top_idx = box_top = None
    for i in range(n - confirm - 1, -1, -1):
        c = h[i]
        if not c or c != c:
            continue
        w = h[i + 1: i + 1 + confirm]
        if len(w) == confirm and all(x < c for x in w):
            box_top_idx, box_top = i, c
            break
    if box_top is None:
        return {"signal": "NO_BOX", "box_top": None, "box_bottom": None,
                "current_price": round(current, decimals)}

    seg = l[box_top_idx:]
    box_bottom = None
    for i in range(len(seg) - confirm):
        c = seg[i]
        if not c or c != c:
            continue
        w = seg[i + 1: i + 1 + confirm]
        if len(w) == confirm and all(x > c for x in w):
            box_bottom = c
            break
    if box_bottom is None:
        valid = [x for x in seg if x and x == x and x > 0]
        box_bottom = min(valid) if valid else None
    if box_bottom is None:
        return {"signal": "NO_BOX", "box_top": round(box_top, decimals),
                "box_bottom": None, "current_price": round(current, decimals)}

    signal = ("BREAKOUT_BUY" if current > box_top else
              "BREAKDOWN_SELL" if current < box_bottom else "IN_BOX")

    rng = box_top - box_bottom
    upside = ((box_top - current) / current * 100) if current else 0
    pos = ((current - box_bottom) / rng * 100) if rng else 0

    return {
        "signal": signal,
        "box_top": round(box_top, decimals),
        "box_bottom": round(box_bottom, decimals),
        "current_price": round(current, decimals),
        "box_range": round(rng, decimals),
        # both historical key spellings — IN/US used the long names,
        # JP/KR/EU the short ones; callers of either keep working
        "upside_pct": round(upside, 2),
        "upside_to_top_pct": round(upside, 2),
        "pos_in_box": round(pos, 1),
        "position_in_box_pct": round(pos, 1),
        "data_points": len(closes),
    }


# ── Shared helpers (verified byte-identical across scans before extraction) ───
def _first_df(ticker, *attrs):
    """First non-empty DataFrame attribute on a yfinance Ticker — NEVER use
    `or` between DataFrames (ambiguous truth value; see yfinance feedback)."""
    for attr in attrs:
        df = getattr(ticker, attr, None)
        if df is not None and isinstance(df, pd.DataFrame) and not df.empty:
            return df
    return None


def _retier(_df, _col):
    """Re-tier liquidity via adaptive_liquidity; never let tiering break a scan."""
    try:
        import adaptive_liquidity as _AL
        return _AL.retier(_df, turnover_col=_col)
    except Exception as _e:
        print(f"  retier skipped: {str(_e)[:60]}")
        return _df


def _bq_fields(df=None, price_round=2):
    """Empty breakout-quality field dict (placeholder when BQ isn't computed)."""
    return {k: None for k in
            ("EMA50", "Above_EMA50", "EMA50_Rising", "Quality_Score",
             "Quality_Grade", "Rel_Volume", "Compression_Ratio", "Body_Pct",
             "Actionable", "Recomputed_Signal")}


def style_sheet(ws):
    """House xlsx sheet styling. Imports openpyxl lazily so scan modules that
    run without it (or headless) never pay the import."""
    try:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    fill_hdr = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    fill_alt = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    font_hdr = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11, bold=False, color="000000")
    thin = Border(
        left=Side(style="thin", color="CBD5E0"), right=Side(style="thin", color="CBD5E0"),
        top=Side(style="thin", color="CBD5E0"), bottom=Side(style="thin", color="CBD5E0"),
    )
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_hdr; cell.fill = fill_hdr; cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_body; cell.border = thin
            if row_idx % 2 == 1:
                cell.fill = fill_alt
            hdr = str(ws.cell(row=1, column=col_idx).value or "").upper()
            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if any(k in hdr for k in ["%", "CAGR", "ROCE", "YIELD"]):
                    cell.number_format = '0.00"%"'
                elif any(k in hdr for k in ["CAP", "LTP", "BOX", "PRICE", "200"]):
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '0.00'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)
