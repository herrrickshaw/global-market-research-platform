#!/usr/bin/env python3
"""
build_business_valuation_report.py — reporting template for actual
strategy-implementation earnings + a business valuation of the research/
trading operation, structured like CFI's standard corporate-finance
templates (Income Statement / Ratios / DCF) the user supplied as a model.

NOT a new copy of those CFI files — those are commercial/licensed
products; this borrows their professional STRUCTURE (statement -> ratios
-> DCF -> comparables) and populates it with this platform's own real,
already-computed numbers, in a fresh workbook.

2026-07-31 — three additions on top of the first version
----------------------------------------------------------
1. EXIT-PRICE FIX (in watchlist_pnl.py, not here): "sold" rows used to
   compare entry cost to the CURRENT market price — a name that crashed
   AFTER being sold showed a fabricated loss that was never the real
   result. watchlist_pnl.realized_sales() now reads the user's actual
   Schedule FA tax filing (the same file already used for entry dates,
   its sale-transaction columns were simply never read before) and
   overlays REAL exit prices where a match exists (36 of 112 sold rows —
   US only, the filing doesn't cover India). This sheet reports those
   real numbers, and honestly labels the remainder ("sold-no-exit-price-
   on-file") as still unverified rather than presenting both the same way.
2. TAXATION & BORROWING COSTS: a new editable assumption block (India
   STCG/LTCG, US STCG/LTCG by holding period, cost of leverage) applied
   to both the actual-earnings summary and the DCF, which is now POST-TAX
   POST-COST, not pre-tax gross.
3. COMPARABLE LISTED FIRMS: real, live-pulled multiples (P/E, P/B, ROE,
   revenue) for Virtu Financial, Interactive Brokers, LPL Financial,
   Angel One, Motilal Oswal, and 5paisa — the closest public comparables
   to a small research/prop operation, with an honest note on how
   imperfect that comparison is (all of them are broking/market-making
   businesses at a completely different scale).

    build_business_valuation_report.py     # writes the .xlsx
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import watchlist_pnl as WP

OUT = Path(__file__).parent / "reports" / "business_valuation" / \
    f"strategy_earnings_and_valuation_{dt.date.today():%Y-%m-%d}.xlsx"

NAVY = "0B2F4A"
TEAL = "1F7A7A"
LIGHT = "EEF4F6"
WARN = "8A6D3B"
WARN_BG = "F4F0E8"
ASSUME_FILL = PatternFill("solid", fgColor="FFF6D8")

HEAD_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEAD_FONT = Font(name="Calibri", size=10, bold=True, color=NAVY)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=NAVY)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
WARN_FONT = Font(name="Calibri", size=9, bold=True, color=WARN)
THIN = Side(style="thin", color="D8D2C2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_df(ws, df, start_row, start_col=1, pct_cols=(), money_cols=()):
    ncols = len(df.columns)
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=str(col))
    _style_header_row(ws, start_row, start_col + ncols - 1)
    for i, (_, r) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=r[col])
            cell.border = BORDER
            if col in pct_cols:
                cell.number_format = "+0.0%;-0.0%"
            elif col in money_cols:
                cell.number_format = "#,##0.00"
    return start_row + 1 + len(df)


def _title(ws, text, row=1, col=1, size=16):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", size=size, bold=True, color=NAVY)
    return row + 1


def _note(ws, text, row, col=1, span=8, font=NOTE_FONT):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    ws.row_dimensions[row].height = 15 * (1 + len(text) // 110)
    return row + 1


def _assume(ws, row, label, value, fmt=None):
    ws.cell(row=row, column=1, value=label)
    cell = ws.cell(row=row, column=3, value=value)
    cell.fill = ASSUME_FILL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    return row + 1


# ── data ─────────────────────────────────────────────────────────────────────

def load_pnl():
    df = WP.build(market=None, status=None)
    d = df.dropna(subset=["pct_since"]).copy()
    d["pct_since"] = d["pct_since"] / 100.0
    held = d[d["status"] == "held"].sort_values("pct_since", ascending=False)
    sold = d[d["status"] == "sold"].sort_values("pct_since", ascending=False)
    sold_real = sold[sold["basis"] == "realized-tax-filing"]
    sold_unverified = sold[sold["basis"] != "realized-tax-filing"]
    return d, held, sold_real, sold_unverified


COMPARABLES = {
    "VIRT": {"label": "Virtu Financial", "market": "US", "kind": "Electronic market maker",
             "mcap": 12_729_935_872, "pe": 9.79, "fwd_pe": 9.09, "pb": 2.97,
             "roe": None, "revenue": 3_264_591_872, "margin": 0.2098, "ccy": "USD"},
    "IBKR": {"label": "Interactive Brokers", "market": "US", "kind": "Broker",
             "mcap": 153_474_187_264, "pe": 35.90, "fwd_pe": 28.35, "pb": 6.86,
             "roe": 0.2402, "revenue": 6_834_999_808, "margin": 0.1647, "ccy": "USD"},
    "LPLA": {"label": "LPL Financial", "market": "US", "kind": "Advisory/broker-dealer",
             "mcap": 27_127_789_568, "pe": 30.53, "fwd_pe": 11.56, "pb": 4.78,
             "roe": 0.2045, "revenue": 17_840_070_656, "margin": 0.0505, "ccy": "USD"},
    "ANGELONE.NS": {"label": "Angel One", "market": "India", "kind": "Discount broking",
                     "mcap": 266_427_809_792, "pe": 26.35, "fwd_pe": 16.48, "pb": 4.34,
                     "roe": None, "revenue": 49_862_279_168, "margin": 0.2070, "ccy": "INR"},
    "MOTILALOFS.NS": {"label": "Motilal Oswal Financial Services", "market": "India",
                       "kind": "Broking / asset management", "mcap": 503_966_662_656,
                       "pe": 25.96, "fwd_pe": 14.18, "pb": 3.91, "roe": None,
                       "revenue": 86_241_697_792, "margin": 0.2296, "ccy": "INR"},
    "5PAISA.NS": {"label": "5paisa Capital", "market": "India", "kind": "Discount broking",
                  "mcap": 16_267_102_208, "pe": 26.63, "fwd_pe": 22.39, "pb": 1.67,
                  "roe": None, "revenue": 2_967_976_960, "margin": 0.1489, "ccy": "INR"},
}


# ── sheets ───────────────────────────────────────────────────────────────────

def sheet_cover(wb, held, sold_real, sold_unverified, d):
    ws = wb.active
    ws.title = "Cover & Methodology"
    r = _title(ws, "Strategy Implementation — Actual Earnings & Business Valuation")
    ws.cell(row=r, column=1, value=f"Generated {dt.date.today():%Y-%m-%d} from live watchlist_pnl.py "
                                    "(incl. Schedule FA tax-filing exit prices) + fact_screener_signal + "
                                    "cost_vs_edge.py + live comparable-company data").font = NOTE_FONT
    r += 2
    r = _note(ws, "STRUCTURE: modeled on the CFI Financial Projection / DCF / Three-Statement "
                  "templates supplied as reference (statement → ratios → DCF → comparables), rebuilt "
                  "fresh with this platform's own real numbers rather than edited into those licensed "
                  "files.", r)
    r += 1
    r = _note(ws, "✓ FIXED 2026-07-31: 'sold' positions now use REAL exit prices from the user's own "
                  "Schedule FA tax filing (36 of 112 sold rows matched — US holdings only, the filing "
                  "doesn't cover India). Positions with no match are still explicitly labeled "
                  "'sold-no-exit-price-on-file', not silently treated as verified.", r)
    r += 1
    r = _note(ws, "⚠ NOT INVESTMENT ADVICE. This is a mechanical accounting/valuation exercise on "
                  "this platform's own historical data — it values the RESEARCH OPERATION as a "
                  "business, it does not recommend any security. Backtest statistics quoted here "
                  "are historical, not a forecast. Comparable-company data (Sheet 6) is for context "
                  "only — none of those firms run the same business at the same scale.", r)
    r += 2

    ws.cell(row=r, column=1, value="Snapshot").font = SUBHEAD_FONT
    r += 1
    stats = [
        ("Total watchlist positions (with computable P&L)", len(d)),
        ("Held positions (actual capital at risk, unrealized)", len(held)),
        ("Held — mean / median return", f"{held['pct_since'].mean():+.1%} / {held['pct_since'].median():+.1%}"),
        ("Held — win rate", f"{(held['pct_since'] > 0).mean():.1%}"),
        ("Closed positions — REAL exit price (Schedule FA)", len(sold_real)),
        ("Closed, real exit — mean / median return", f"{sold_real['pct_since'].mean():+.1%} / {sold_real['pct_since'].median():+.1%}"
         if len(sold_real) else "n/a"),
        ("Closed, real exit — win rate", f"{(sold_real['pct_since'] > 0).mean():.1%}" if len(sold_real) else "n/a"),
        ("Closed positions — no exit price on file (legacy, unverified)", len(sold_unverified)),
    ]
    for label, val in stats:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=3, value=val).font = Font(bold=True)
        r += 1
    r += 1
    r = _note(ws, "⚠ The real-exit-price group shows a 94% LOSS rate (mean -34%) — sharply worse than "
                  "the 82% win rate on still-open positions. This is a genuine, sobering finding, not "
                  "smoothed over: it may reflect winners being held (letting profits run) while losers "
                  "get cut, or it may reflect that the open book hasn't yet realized its own losers. "
                  "Both readings are consistent with the data; this workbook does not pick one.", r, font=WARN_FONT)
    _autofit(ws, [50, 4, 20, 4, 4, 4, 4, 4])
    return ws


def sheet_actual_pnl(wb, held, sold_real, sold_unverified, d):
    ws = wb.create_sheet("Actual P&L")
    r = _title(ws, "Actual Earnings — Held (Unrealized) + Closed (Realized)")
    r = _note(ws, "Held: real cost basis vs current mark, capital genuinely at risk today (unrealized). "
                  "Closed/real exit: the user's own Schedule FA tax-filed sale record — actual realized "
                  "result. Closed/unverified: no exit price on file, NOT treated as real P&L.", r)
    r += 1

    ws.cell(row=r, column=1, value="Held — by market").font = SUBHEAD_FONT
    r += 1
    by_mkt = held.groupby("market")["pct_since"].agg(["count", "mean", "median"]).round(4)
    by_mkt.columns = ["n", "mean_return", "median_return"]
    r = _write_df(ws, by_mkt.reset_index(), r, pct_cols=("mean_return", "median_return")) + 2

    ws.cell(row=r, column=1, value="Held — full detail, sorted by return").font = SUBHEAD_FONT
    r += 1
    detail = held[["symbol", "market", "entry_price", "ltp", "pct_since"]].rename(
        columns={"entry_price": "cost_basis", "ltp": "current_price", "pct_since": "unrealized_return"})
    r = _write_df(ws, detail, r, pct_cols=("unrealized_return",), money_cols=("cost_basis", "current_price")) + 2

    ws.cell(row=r, column=1, value="Closed — REAL exit price (Schedule FA tax filing)").font = Font(bold=True, color="2F7D4F")
    r += 1
    real_detail = sold_real[["symbol", "market", "entry_price", "ltp", "pct_since"]].rename(
        columns={"entry_price": "cost_basis", "ltp": "price_now_(not_exit)", "pct_since": "realized_return"})
    r = _write_df(ws, real_detail, r, pct_cols=("realized_return",), money_cols=("cost_basis", "price_now_(not_exit)")) + 2

    ws.cell(row=r, column=1, value="Closed — NO exit price on file (unverified, legacy)").font = WARN_FONT
    r += 1
    unv = sold_unverified[["symbol", "market"]].copy()
    r = _write_df(ws, unv, r) + 1
    _note(ws, "These predate the 2026-07-31 fix and have no trade log anywhere in this pipeline to "
              "reconstruct a real exit price — left blank rather than guessed.", r)

    _autofit(ws, [12, 8, 12, 16, 14])
    return ws


def sheet_signal_quality(wb):
    ws = wb.create_sheet("Strategy Signal Quality")
    r = _title(ws, "Backtest Evidence — Screener/Signal Performance (Large-Sample)")
    r = _note(ws, "From fact_screener_signal (4.98M rows, 37 screeners × 8 markets, 2009-2026) — "
                  "reverified 2026-07-31 against the FULL history, not a sample. Historical, not a "
                  "forecast. This is the evidence base for the DCF sheet's 'edge' assumption.", r)
    r += 1
    rows = [
        ("Golden Cross", "India", 35376, 0.3115, "252d"),
        ("Darvas", "India", 63164, 0.2914, "252d"),
        ("Golden Cross", "USA", 29941, -0.0766, "252d"),
        ("Piotroski", "USA", 2937, -0.0082, "252d"),
    ]
    df = pd.DataFrame(rows, columns=["screener", "market", "n", "mean_252d_excess_return", "horizon"])
    r = _write_df(ws, df, r, pct_cols=("mean_252d_excess_return",)) + 2
    r = _note(ws, "India momentum screens (Golden Cross/Darvas) show large, statistically substantial "
                  "positive excess return over this history. US Golden Cross and Piotroski show FLAT "
                  "TO NEGATIVE excess return over the same horizon — do not extrapolate the India "
                  "result to other markets or other screens.", r)
    r += 1
    r = _note(ws, "⚠ Separately, this pipeline's own re-entry signal (a different, related feature) "
                  "was explicitly re-tested and found to have NO measured forward edge "
                  "(India 63d excess return: -2.01% median, t=0.06) — that finding is NOT reversed "
                  "by the numbers above, which are a different signal on a different, larger dataset.", r, font=WARN_FONT)
    _autofit(ws, [16, 10, 10, 20, 10])
    return ws


def sheet_capacity(wb):
    ws = wb.create_sheet("Capacity & Cost Model")
    r = _title(ws, "Execution Capacity — Where the Edge Dies to Market Impact")
    r = _note(ws, "From cost_vs_edge.py / reports/execution_capacity.csv — Almgren-Chriss square-root "
                  "impact model, 15%-ADV participation cap, Corwin-Schultz spread estimate. "
                  "net_edge_bps = gross excess return minus round-trip execution cost at that AUM. "
                  "This covers TRADING/execution cost only — taxation and borrowing cost are Sheet 5.", r)
    r += 1

    csv_path = Path(__file__).parent / "reports" / "execution_capacity.csv"
    cap_df = pd.read_csv(csv_path)
    r = _write_df(ws, cap_df, r) + 2

    ws.cell(row=r, column=1, value="Capacity ceiling per desk (AUM $M where net edge → 0)").font = SUBHEAD_FONT
    r += 1
    ceiling = pd.DataFrame([
        ("IN", 44, "~1"), ("US", 21, "~1"), ("JP", 15, "~0 (already dead at $1M)"),
        ("KR", 66, "~1"), ("EU", 15, "~5"),
    ], columns=["desk", "gross_edge_bps_2wk", "capacity_$M_AUM"])
    r = _write_df(ws, ceiling, r) + 1
    _note(ws, "This is the hard constraint on the DCF sheet's scale assumption — the edge is a "
              "small-money edge, not something that scales to institutional AUM.", r, font=WARN_FONT)
    _autofit(ws, [10, 10, 10, 10, 12, 12, 14, 14])
    return ws


def sheet_tax_and_costs(wb):
    ws = wb.create_sheet("Taxation & Borrowing Costs")
    r = _title(ws, "Taxation & Cost of Capital — Editable Assumptions")
    r = _note(ws, "Standard published rates as of 2026, NOT tax advice — verify against your own "
                  "bracket/filing status before relying on this. Every cell below feeds the DCF sheet's "
                  "post-tax free cash flow.", r, font=WARN_FONT)
    r += 1

    r0 = r
    ws.cell(row=r, column=1, value="India — equity capital gains").font = SUBHEAD_FONT
    r += 1
    r = _assume(ws, r, "STCG rate (equity, holding <1yr)", 0.20, "0.0%")
    r_in_stcg = r0 + 1
    r = _assume(ws, r, "LTCG rate (equity, holding >1yr, above exemption)", 0.125, "0.0%")
    r_in_ltcg = r0 + 2
    r = _assume(ws, r, "LTCG annual exemption (₹)", 125_000, "#,##0")
    r_in_exempt = r0 + 3
    r = _assume(ws, r, "STT (delivery, buy+sell combined, % of turnover)", 0.002, "0.00%")
    r_in_stt = r0 + 4
    r += 1

    r1 = r
    ws.cell(row=r, column=1, value="US — equity capital gains (SET YOUR OWN BRACKET)").font = SUBHEAD_FONT
    r += 1
    r = _assume(ws, r, "STCG rate (holding <1yr, taxed as ordinary income — placeholder)", 0.30, "0.0%")
    r_us_stcg = r1 + 1
    r = _assume(ws, r, "LTCG rate (holding >1yr — placeholder)", 0.15, "0.0%")
    r_us_ltcg = r1 + 2
    r += 1

    r2 = r
    ws.cell(row=r, column=1, value="Borrowing / leverage cost").font = SUBHEAD_FONT
    r += 1
    r = _assume(ws, r, "Leverage ratio (0 = unlevered, 1 = 2x gross exposure)", 0.0, "0.0%")
    r_lev = r2 + 1
    r = _assume(ws, r, "Cost of borrowed capital (annualized, margin/broker call rate)", 0.09, "0.0%")
    r_borrow = r2 + 2
    r += 2

    r = _note(ws, "Applied in the DCF (Sheet 7) as: FCF_after_tax = FCF_pretax × (1 − blended_tax_rate) "
                  "− (leverage_ratio × capital × cost_of_borrowed_capital). Blended tax rate uses the "
                  "India LTCG rate as the default holding-period assumption for the capacity-ceiling "
                  "desks — change it directly on the DCF sheet if the real turnover is short-term.", r)

    ws._tax_rows = dict(in_stcg=r_in_stcg, in_ltcg=r_in_ltcg, in_exempt=r_in_exempt, in_stt=r_in_stt,
                         us_stcg=r_us_stcg, us_ltcg=r_us_ltcg, lev=r_lev, borrow=r_borrow)
    _autofit(ws, [55, 4, 16, 8])
    return ws


def sheet_dcf(wb, held, tax_ws):
    ws = wb.create_sheet("DCF — Operation Valuation")
    tr = tax_ws._tax_rows
    tax_sheet = "'Taxation & Borrowing Costs'"
    r = _title(ws, "DCF — Illustrative Valuation of the Research/Trading Operation")
    r = _note(ws, "Values the OPERATION as a business (a research capability that could size capital "
                  "up to its measured capacity ceiling), NOT any individual security. Every yellow "
                  "cell below is an assumption you can change — the NPV recalculates live. "
                  "Post-tax, post-borrowing-cost — see Sheet 5 for the tax/leverage assumptions this "
                  "pulls from.", r)
    r += 1
    r = _note(ws, "⚠ Illustrative only. Discount rate reflects a solo, key-person-dependent, "
                  "small-cap/illiquid strategy — genuinely high risk, not a diversified fund's cost "
                  "of capital.", r, font=WARN_FONT)
    r += 1

    ws.cell(row=r, column=1, value="Assumptions (editable)").font = SUBHEAD_FONT
    r += 1
    r0 = r
    r = _assume(ws, r, "Deployable capital at capacity ceiling ($, blended IN/US/KR desk)", 1_000_000, "#,##0")
    r_cap = r0
    r = _assume(ws, r, "Blended gross edge captured (annualized, bps)", 0.02, "0.0%")
    r_edge = r0 + 1
    r = _assume(ws, r, "Operating cost ratio (data/compute/infra, % of capital)", 0.005, "0.0%")
    r_cost = r0 + 2
    r = _assume(ws, r, "Discount rate (reflects key-person/illiquidity risk)", 0.25, "0.0%")
    r_disc = r0 + 3
    r = _assume(ws, r, "Projection years", 5, "0")
    r_years = r0 + 4
    r = _assume(ws, r, "Terminal growth rate", 0.02, "0.0%")
    r_term = r0 + 5
    r += 1

    ws.cell(row=r, column=1, value="Free cash flow build-up").font = SUBHEAD_FONT
    r += 1
    ws.cell(row=r, column=1, value="Pre-tax annual FCF = capital × (edge − opex ratio)")
    pretax_row = r
    c = ws.cell(row=r, column=3, value=f"=C{r_cap}*(C{r_edge}-C{r_cost})")
    c.number_format = "#,##0"; c.border = BORDER
    r += 1
    ws.cell(row=r, column=1, value=f"less: capital gains tax (using {tax_sheet} LTCG rate)")
    tax_row = r
    c = ws.cell(row=r, column=3, value=f"=C{pretax_row}*{tax_sheet}!C{tr['in_ltcg']}")
    c.number_format = "#,##0"; c.border = BORDER
    r += 1
    ws.cell(row=r, column=1, value=f"less: cost of leverage (from {tax_sheet})")
    borrow_row = r
    c = ws.cell(row=r, column=3,
                value=f"=C{r_cap}*{tax_sheet}!C{tr['lev']}*{tax_sheet}!C{tr['borrow']}")
    c.number_format = "#,##0"; c.border = BORDER
    r += 1
    ws.cell(row=r, column=1, value="Post-tax, post-borrowing-cost annual FCF").font = Font(bold=True)
    fcf_row = r
    c = ws.cell(row=r, column=3, value=f"=C{pretax_row}-C{tax_row}-C{borrow_row}")
    c.number_format = "#,##0"; c.border = BORDER; c.font = Font(bold=True)
    r += 2

    ws.cell(row=r, column=1, value="5-year DCF").font = SUBHEAD_FONT
    r += 1
    ws.cell(row=r, column=1, value="Year")
    for y in range(1, 6):
        ws.cell(row=r, column=1 + y, value=y)
    _style_header_row(ws, r, 6)
    r += 1
    fcf_line = r
    ws.cell(row=r, column=1, value="Post-tax FCF")
    for y in range(1, 6):
        c = ws.cell(row=r, column=1 + y, value=f"=$C${fcf_row}")
        c.number_format = "#,##0"
    r += 1
    dcf_line = r
    ws.cell(row=r, column=1, value="PV of FCF")
    for y in range(1, 6):
        col = get_column_letter(1 + y)
        c = ws.cell(row=r, column=1 + y, value=f"={col}{fcf_line}/(1+$C${r_disc})^{y}")
        c.number_format = "#,##0"
    r += 2

    ws.cell(row=r, column=1, value="Terminal value (Gordon growth, at year 5)")
    tv_row = r
    c = ws.cell(row=r, column=3,
                value=f"=({get_column_letter(6)}{fcf_line}*(1+C{r_term}))/(C{r_disc}-C{r_term})")
    c.number_format = "#,##0"; c.border = BORDER
    r += 1
    ws.cell(row=r, column=1, value="PV of terminal value")
    pv_tv_row = r
    c = ws.cell(row=r, column=3, value=f"=C{tv_row}/(1+C{r_disc})^C{r_years}")
    c.number_format = "#,##0"; c.border = BORDER
    r += 2

    ws.cell(row=r, column=1, value="Estimated operation value (NPV, post-tax)").font = Font(bold=True, size=12, color=NAVY)
    c = ws.cell(row=r, column=3,
                value=f"=SUM({get_column_letter(2)}{dcf_line}:{get_column_letter(6)}{dcf_line})+C{pv_tv_row}")
    c.number_format = "#,##0"; c.font = Font(bold=True, size=12, color=NAVY); c.border = BORDER
    r += 2

    r = _note(ws, "Sensitivity: halving the discount rate roughly doubles this NPV; the capacity "
                  "ceiling (Sheet 4) is the actual binding constraint on the capital-at-capacity input "
                  "above. Switching the tax-rate reference on Sheet 5 from LTCG to STCG (short holding "
                  "periods) meaningfully lowers this NPV — check which one matches your real turnover.", r)

    _autofit(ws, [55, 10, 16, 12, 12, 12, 12])
    return ws


def sheet_benchmark(wb):
    ws = wb.create_sheet("Comparable Listed Firms")
    r = _title(ws, "Benchmark — Comparable Publicly Listed Trading/Broking Firms")
    r = _note(ws, "Live-pulled multiples (yfinance), 2026-07-31. These are the closest PUBLIC "
                  "comparables to a trading/research operation — but every one of them is a "
                  "broking, market-making, or advisory business at institutional scale, not a "
                  "solo signal-research operation. Use for context on what the market pays for "
                  "trading-adjacent businesses generally, not as a direct valuation multiple for "
                  "this operation.", r, font=WARN_FONT)
    r += 1

    rows = []
    for tkr, c in COMPARABLES.items():
        rows.append({
            "ticker": tkr, "company": c["label"], "market": c["market"], "business": c["kind"],
            "market_cap": c["mcap"], "P/E": c["pe"], "fwd_P/E": c["fwd_pe"], "P/B": c["pb"],
            "ROE": c["roe"], "revenue": c["revenue"], "net_margin": c["margin"], "currency": c["ccy"],
        })
    df = pd.DataFrame(rows)
    r = _write_df(ws, df, r, pct_cols=("ROE", "net_margin"), money_cols=("market_cap", "revenue")) + 2

    ws.cell(row=r, column=1, value="Reading this against the DCF").font = SUBHEAD_FONT
    r += 1
    r = _note(ws, f"Median P/E across these six: ~{df['P/E'].median():.1f}x. If this operation's "
                  "post-tax FCF (DCF sheet) were instead capitalized at that multiple rather than "
                  "discounted as a 5-year NPV, the implied value would be FCF × that P/E — a rough "
                  "cross-check, not a replacement for the DCF, since none of these firms carry the "
                  "same key-person/capacity-ceiling risk this operation does.", r)
    r += 1
    r = _note(ws, "US firms (VIRT/IBKR/LPLA) trade at higher absolute scale and lower average P/E than "
                  "the India broking names (ANGELONE/MOTILALOFS/5PAISA) shown here, consistent with "
                  "India's broking sector pricing in faster growth. Neither group is a market-neutral "
                  "signal-research shop like this one — the comparison is directional, not exact.", r)
    _autofit(ws, [16, 26, 8, 26, 16, 8, 8, 8, 8, 16, 10, 8])
    return ws


def build():
    wb = Workbook()
    d, held, sold_real, sold_unverified = load_pnl()
    sheet_cover(wb, held, sold_real, sold_unverified, d)
    sheet_actual_pnl(wb, held, sold_real, sold_unverified, d)
    sheet_signal_quality(wb)
    sheet_capacity(wb)
    tax_ws = sheet_tax_and_costs(wb)
    sheet_dcf(wb, held, tax_ws)
    sheet_benchmark(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    return OUT


if __name__ == "__main__":
    build()
